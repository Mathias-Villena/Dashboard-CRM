import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Paleta de colores Premium
COLOR_TEAL = "0F766E"      # #0f766e (Color primario)
COLOR_SLATE_50 = "F8FAFC"  # #f8fafc
COLOR_SLATE_100 = "F1F5F9" # #f1f5f9
COLOR_SLATE_200 = "E2E8F0" # #e2e8f0
COLOR_TEXT_MAIN = "0F172A" # #0f172a
COLOR_WHITE = "FFFFFF"

def get_thin_border():
    thin_side = Side(border_style="thin", color="CBD5E1")
    return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def generate_excel_report(df_opportunities):
    """
    Genera un archivo Excel (.xlsx) altamente estilizado con openpyxl.
    Devuelve los bytes del archivo generado listos para descarga.
    """
    wb = Workbook()
    
    # 1. Pestaña de Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"
    ws_resumen.views.sheetView[0].showGridLines = True
    
    # Fuentes y estilos
    font_title = Font(name="Segoe UI", size=16, bold=True, color=COLOR_TEAL)
    font_section = Font(name="Segoe UI", size=12, bold=True, color=COLOR_TEXT_MAIN)
    font_header = Font(name="Segoe UI", size=11, bold=True, color=COLOR_WHITE)
    font_data = Font(name="Segoe UI", size=10, color=COLOR_TEXT_MAIN)
    font_bold = Font(name="Segoe UI", size=10, bold=True, color=COLOR_TEXT_MAIN)
    
    fill_header = PatternFill(start_color=COLOR_TEAL, end_color=COLOR_TEAL, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_SLATE_50, end_color=COLOR_SLATE_50, fill_type="solid")
    fill_total = PatternFill(start_color=COLOR_SLATE_100, end_color=COLOR_SLATE_100, fill_type="solid")
    
    border_thin = get_thin_border()
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Título
    ws_resumen["B2"] = "REPORTE COMERCIAL CRM - GOHIGHLEVEL"
    ws_resumen["B2"].font = font_title
    ws_resumen["B3"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_resumen["B3"].font = Font(name="Segoe UI", size=10, italic=True)
    
    # --- Tabla 1: Métricas de Etapas ---
    ws_resumen["B5"] = "Oportunidades por Etapa"
    ws_resumen["B5"].font = font_section
    
    headers_etapas = ["Categoría / Etapa", "Cantidad", "Porcentaje (%)"]
    for col_idx, h in enumerate(headers_etapas, start=2):
        cell = ws_resumen.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    # Agrupaciones
    # Mapeo de etapas
    stages_mapping = {
        "fdc27149-b398-4ed7-9271-946c66dc9f0f": "EN APROBACION WIN",
        "085ad64f-769b-42f7-986d-6eef802f0634": "PENDIENTE CARTA",
        "f251b78c-b57f-4cbd-aa61-3653b54c7677": "EN CONSTRUCCION",
        "46400c15-10a3-4c96-a5b0-76f0cf65b753": "EN CONSTRUCCION",
        "dc5a218f-50a8-4bb6-9351-82b2f10d9886": "PROYECTO LIQUIDADO"
    }
    
    total_opps = len(df_opportunities)
    
    def get_stage_cat(stage_id):
        return stages_mapping.get(stage_id, "OTROS")
        
    df_stages = df_opportunities.copy()
    if not df_stages.empty:
        df_stages["categoria"] = df_stages["pipelineStageId"].apply(get_stage_cat)
        counts = df_stages["categoria"].value_counts().to_dict()
    else:
        counts = {}
        
    categorias_ordenadas = ["EN APROBACION WIN", "PENDIENTE CARTA", "EN CONSTRUCCION", "PROYECTO LIQUIDADO", "OTROS"]
    row_idx = 7
    for cat in categorias_ordenadas:
        qty = counts.get(cat, 0)
        pct = (qty / total_opps) * 100 if total_opps > 0 else 0
        
        # Escribir fila
        c1 = ws_resumen.cell(row=row_idx, column=2, value=cat)
        c2 = ws_resumen.cell(row=row_idx, column=3, value=qty)
        c3 = ws_resumen.cell(row=row_idx, column=4, value=pct / 100)
        
        c1.font = font_data
        c1.border = border_thin
        c1.alignment = align_left
        
        c2.font = font_data
        c2.border = border_thin
        c2.alignment = align_right
        c2.number_format = "#,##0"
        
        c3.font = font_data
        c3.border = border_thin
        c3.alignment = align_right
        c3.number_format = "0.0%"
        
        if row_idx % 2 == 1:
            c1.fill = fill_zebra
            c2.fill = fill_zebra
            c3.fill = fill_zebra
            
        row_idx += 1
        
    # Fila de Total
    c1 = ws_resumen.cell(row=row_idx, column=2, value="TOTAL OPORTUNIDADES")
    c2 = ws_resumen.cell(row=row_idx, column=3, value=total_opps)
    c3 = ws_resumen.cell(row=row_idx, column=4, value=1.0)
    for c in [c1, c2, c3]:
        c.font = font_bold
        c.fill = fill_total
        c.border = border_thin
    c1.alignment = align_left
    c2.alignment = align_right
    c2.number_format = "#,##0"
    c3.alignment = align_right
    c3.number_format = "0.0%"
    
    # --- Tabla 2: Oportunidades por Gestor ---
    start_row_gestores = row_idx + 3
    ws_resumen.cell(row=start_row_gestores, column=2, value="Resumen por Gestor").font = font_section
    
    headers_gestores = ["Gestor (Hunter)", "Cantidad", "Porcentaje (%)"]
    for col_idx, h in enumerate(headers_gestores, start=2):
        cell = ws_resumen.cell(row=start_row_gestores+1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    if not df_opportunities.empty:
        df_gestores = df_opportunities["gestor"].value_counts().to_dict()
    else:
        df_gestores = {}
        
    row_idx = start_row_gestores + 2
    for gestor, qty in df_gestores.items():
        pct = (qty / total_opps) * 100 if total_opps > 0 else 0
        
        c1 = ws_resumen.cell(row=row_idx, column=2, value=gestor)
        c2 = ws_resumen.cell(row=row_idx, column=3, value=qty)
        c3 = ws_resumen.cell(row=row_idx, column=4, value=pct / 100)
        
        c1.font = font_data
        c1.border = border_thin
        c1.alignment = align_left
        
        c2.font = font_data
        c2.border = border_thin
        c2.alignment = align_right
        c2.number_format = "#,##0"
        
        c3.font = font_data
        c3.border = border_thin
        c3.alignment = align_right
        c3.number_format = "0.0%"
        
        if row_idx % 2 == 1:
            c1.fill = fill_zebra
            c2.fill = fill_zebra
            c3.fill = fill_zebra
        row_idx += 1
        
    # Fila de Total Gestores
    c1 = ws_resumen.cell(row=row_idx, column=2, value="TOTAL GESTORES")
    c2 = ws_resumen.cell(row=row_idx, column=3, value=total_opps)
    c3 = ws_resumen.cell(row=row_idx, column=4, value=1.0)
    for c in [c1, c2, c3]:
        c.font = font_bold
        c.fill = fill_total
        c.border = border_thin
    c1.alignment = align_left
    c2.alignment = align_right
    c2.number_format = "#,##0"
    c3.alignment = align_right
    c3.number_format = "0.0%"
    
    # 2. Pestaña de Detalle Completo
    ws_detalle = wb.create_sheet(title="Listado Detallado")
    ws_detalle.views.sheetView[0].showGridLines = True
    
    headers_detalle = [
        "ID Oportunidad", "Nombre Oportunidad", "Cliente / Contacto", 
        "Gestor (Hunter)", "Distrito", "Etapa Pipeline", "Estado GHL", 
        "Valor Monetario", "Fecha Creación", "Email Contacto", "Teléfono Contacto"
    ]
    
    # Escribir Headers
    for col_idx, h in enumerate(headers_detalle, start=1):
        cell = ws_detalle.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    # Escribir filas de datos
    for r_idx, opp in enumerate(df_opportunities.to_dict(orient="records"), start=2):
        created_date_str = opp.get("createdAt")
        if created_date_str:
            try:
                # Tratar de formatear la fecha
                dt = datetime.fromisoformat(created_date_str.replace("Z", "+00:00"))
                fecha_formato = dt.strftime("%d/%m/%Y")
            except:
                fecha_formato = str(created_date_str)
        else:
            fecha_formato = "-"
            
        stage_id = opp.get("pipelineStageId")
        stage_name = stages_mapping.get(stage_id, f"Otro ({stage_id[:8]}...)")
        
        ws_detalle.cell(row=r_idx, column=1, value=opp.get("id")).alignment = align_center
        ws_detalle.cell(row=r_idx, column=2, value=opp.get("name")).alignment = align_left
        ws_detalle.cell(row=r_idx, column=3, value=opp.get("contactName")).alignment = align_left
        ws_detalle.cell(row=r_idx, column=4, value=opp.get("gestor")).alignment = align_center
        ws_detalle.cell(row=r_idx, column=5, value=opp.get("distrito")).alignment = align_left
        ws_detalle.cell(row=r_idx, column=6, value=stage_name).alignment = align_left
        ws_detalle.cell(row=r_idx, column=7, value=opp.get("status")).alignment = align_center
        
        cell_val = ws_detalle.cell(row=r_idx, column=8, value=opp.get("monetaryValue") or 0.0)
        cell_val.alignment = align_right
        cell_val.number_format = "$#,##0.00"
        
        ws_detalle.cell(row=r_idx, column=9, value=fecha_formato).alignment = align_center
        ws_detalle.cell(row=r_idx, column=10, value=opp.get("contactEmail")).alignment = align_left
        ws_detalle.cell(row=r_idx, column=11, value=opp.get("contactPhone")).alignment = align_left
        
        # Aplicar fuente, borde y cebra
        for col_idx in range(1, 12):
            cell = ws_detalle.cell(row=r_idx, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            if r_idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Autoajustar columnas en ambas hojas
    for ws in [ws_resumen, ws_detalle]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    # Guardar en un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


class ExecutiveReportPDF(FPDF):
    """
    Clase FPDF personalizada para generar un reporte PDF premium y limpio.
    """
    def header(self):
        # Banner superior en color Teal
        self.set_fill_color(15, 118, 110) # #0f766e
        self.rect(0, 0, 210, 32, 'F')
        
        self.set_text_color(255, 255, 255)
        self.set_y(6)
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 8, "REPORTE COMERCIAL CRM - GOHIGHLEVEL", ln=True, align="C")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 5, f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Confidencial", ln=True, align="C")
        self.ln(15)
        
    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(148, 163, 184) # Slate-400
        # Línea divisoria
        self.set_draw_color(226, 232, 240)
        self.line(10, 280, 200, 280)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="R")
        self.set_x(10)
        self.cell(0, 10, "Romol Capital - CRM Dashboard", align="L")

def generate_pdf_report(df_opportunities, summary_metrics):
    """
    Genera un reporte ejecutivo en formato PDF estilizado.
    """
    pdf = ExecutiveReportPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(15, 20, 15)
    
    # 1. TÍTULO DE SECCIÓN Y RESUMEN KPI
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(15, 118, 110) # Teal-700
    pdf.cell(0, 10, "1. KPI CENTRALES DEL NEGOCIO", ln=True)
    pdf.ln(2)
    
    # Dibujar tarjetas de métricas en formato tabla/celdas
    pdf.set_fill_color(248, 250, 252) # Slate-50
    pdf.set_draw_color(226, 232, 240) # Slate-200
    pdf.set_text_color(15, 23, 42) # Slate-900
    
    col_w = 45
    h_card = 18
    
    # KPIs en formato fila
    kpis = [
        {"title": "Total Oportunidades", "value": str(summary_metrics.get("total", 0))},
        {"title": "En Aprobación WIN", "value": str(summary_metrics.get("en_aprobacion", 0))},
        {"title": "Pendiente Carta", "value": str(summary_metrics.get("pendiente_carta", 0))},
        {"title": "En Construcción", "value": str(summary_metrics.get("en_construccion", 0))},
    ]
    
    for i, kpi in enumerate(kpis):
        # Dibujar fondo y borde
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.rect(x, y, col_w, h_card, 'FD')
        
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(100, 116, 139) # Slate-500
        pdf.set_xy(x, y + 2)
        pdf.cell(col_w, 4, kpi["title"], align="C", ln=True)
        
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(15, 118, 110) # Teal-700
        pdf.set_x(x)
        pdf.cell(col_w, 8, kpi["value"], align="C")
        
        # Mover a la derecha para la siguiente tarjeta
        pdf.set_xy(x + col_w + 3, y)
        
    pdf.ln(h_card + 8)
    
    # 2. RESUMEN GESTORES Y DISTRITOS (TABLAS LADO A LADO O SEGUIDAS)
    pdf.set_x(15)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(15, 118, 110) # Teal-700
    pdf.cell(0, 10, "2. DISTRIBUCION COMERCIAL", ln=True)
    pdf.ln(2)
    
    total_opps = len(df_opportunities)
    
    # Tabla Gestores
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(15, 118, 110)
    pdf.set_text_color(255, 255, 255)
    
    # Anchos de columna
    w_g1, w_g2, w_g3 = 60, 40, 40
    
    pdf.cell(w_g1, 7, "Gestor (Hunter)", border=1, fill=True, align="L")
    pdf.cell(w_g2, 7, "Oportunidades", border=1, fill=True, align="C")
    pdf.cell(w_g3, 7, "Porcentaje", border=1, fill=True, align="C")
    pdf.ln()
    
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(15, 23, 42)
    
    if not df_opportunities.empty:
        counts_gestores = df_opportunities["gestor"].value_counts().to_dict()
    else:
        counts_gestores = {}
        
    for idx, (gestor, qty) in enumerate(counts_gestores.items()):
        pct = (qty / total_opps) * 100 if total_opps > 0 else 0
        
        # Cebra
        fill = (idx % 2 == 1)
        pdf.set_fill_color(248, 250, 252) if fill else pdf.set_fill_color(255, 255, 255)
        
        pdf.cell(w_g1, 6, f"  {gestor}", border=1, fill=fill, align="L")
        pdf.cell(w_g2, 6, str(qty), border=1, fill=fill, align="C")
        pdf.cell(w_g3, 6, f"{pct:.1f}%", border=1, fill=fill, align="C")
        pdf.ln()
        
    pdf.ln(6)
    
    # Tabla Distritos (Top 5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 8, "Top 5 Distritos con Mayor Oportunidades", ln=True)
    
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(15, 118, 110)
    pdf.set_text_color(255, 255, 255)
    
    pdf.cell(w_g1, 7, "Distrito", border=1, fill=True, align="L")
    pdf.cell(w_g2, 7, "Oportunidades", border=1, fill=True, align="C")
    pdf.cell(w_g3, 7, "Porcentaje", border=1, fill=True, align="C")
    pdf.ln()
    
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(15, 23, 42)
    
    if not df_opportunities.empty:
        counts_distritos = df_opportunities["distrito"].value_counts().head(5).to_dict()
    else:
        counts_distritos = {}
        
    for idx, (distrito, qty) in enumerate(counts_distritos.items()):
        pct = (qty / total_opps) * 100 if total_opps > 0 else 0
        
        # Cebra
        fill = (idx % 2 == 1)
        pdf.set_fill_color(248, 250, 252) if fill else pdf.set_fill_color(255, 255, 255)
        
        pdf.cell(w_g1, 6, f"  {distrito}", border=1, fill=fill, align="L")
        pdf.cell(w_g2, 6, str(qty), border=1, fill=fill, align="C")
        pdf.cell(w_g3, 6, f"{pct:.1f}%", border=1, fill=fill, align="C")
        pdf.ln()
        
    # 3. LISTADO DETALLADO EN LA SIGUIENTE PÁGINA
    pdf.add_page()
    
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 10, "3. DETALLE DE OPORTUNIDADES ACTIVAS", ln=True)
    pdf.ln(2)
    
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(15, 118, 110)
    pdf.set_text_color(255, 255, 255)
    
    # Anchos de columna en detalle
    w_d1, w_d2, w_d3, w_d4, w_d5 = 65, 30, 35, 25, 25
    pdf.cell(w_d1, 6, "Nombre del Proyecto / Oportunidad", border=1, fill=True, align="L")
    pdf.cell(w_d2, 6, "Gestor", border=1, fill=True, align="C")
    pdf.cell(w_d3, 6, "Distrito", border=1, fill=True, align="C")
    pdf.cell(w_d4, 6, "Estado GHL", border=1, fill=True, align="C")
    pdf.cell(w_d5, 6, "Valor", border=1, fill=True, align="R")
    pdf.ln()
    
    pdf.set_font("Helvetica", "", 7.5)
    pdf.set_text_color(15, 23, 42)
    
    # Escribir filas del listado detallado (limitado a los primeros 35 para que entre limpio en la hoja)
    lista_ordenada = df_opportunities.sort_values(by="createdAt", ascending=False)
    for idx, opp in enumerate(lista_ordenada.to_dict(orient="records")):
        # Controlar cebra
        fill = (idx % 2 == 1)
        pdf.set_fill_color(248, 250, 252) if fill else pdf.set_fill_color(255, 255, 255)
        
        # Limitar longitud de texto para que no desborde la celda
        opp_name = opp.get("name") or "Sin Nombre"
        if len(opp_name) > 38:
            opp_name = opp_name[:35] + "..."
            
        distrito_name = opp.get("distrito") or "SIN DISTRITO"
        if len(distrito_name) > 18:
            distrito_name = distrito_name[:16] + "..."
            
        pdf.cell(w_d1, 5.5, f" {opp_name}", border=1, fill=fill, align="L")
        pdf.cell(w_d2, 5.5, opp.get("gestor", "OTROS"), border=1, fill=fill, align="C")
        pdf.cell(w_d3, 5.5, distrito_name, border=1, fill=fill, align="C")
        pdf.cell(w_d4, 5.5, opp.get("status", "open").upper(), border=1, fill=fill, align="C")
        
        monetary_val = opp.get("monetaryValue") or 0.0
        pdf.cell(w_d5, 5.5, f"${monetary_val:,.1f}  ", border=1, fill=fill, align="R")
        pdf.ln()
        
        # Si la lista es muy larga, el paginado automático de FPDF2 creará nuevas hojas
        
    return pdf.output()
